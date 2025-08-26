import pandas as pd
import os
import sys
import json
import base64
import requests
from datetime import datetime
import streamlit as st

# Add current directory to Python path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from crypto_utils import DataEncryption
from config import ConfigManager, POINTS_CORRECT_RESULT, POINTS_EXACT_SCORE, POINTS_GOAL_DIFFERENCE

class GitHubDataManager:
    def __init__(self):
        # Initialize encryption with Streamlit secrets
        self.encryption = self._initialize_encryption()
        self.config = ConfigManager()
        
        # GitHub configuration - get from Streamlit secrets or environment variables
        self.github_token = self._get_secret('GITHUB_TOKEN')
        self.repo_owner = self._get_secret('GITHUB_REPO_OWNER')
        self.repo_name = self._get_secret('GITHUB_REPO_NAME')
        self.branch = 'main'
        
        if not all([self.github_token, self.repo_owner, self.repo_name]):
            raise ValueError("GitHub configuration missing. Set GITHUB_TOKEN, GITHUB_REPO_OWNER, and GITHUB_REPO_NAME in Streamlit secrets or environment variables.")
        
        self.base_url = f"https://api.github.com/repos/{self.repo_owner}/{self.repo_name}/contents"
        self.headers = {
            'Authorization': f'token {self.github_token}',
            'Accept': 'application/vnd.github.v3+json'
        }
    
    def _get_secret(self, key):
        """Get secret from Streamlit secrets or environment variables"""
        try:
            # Try Streamlit secrets first
            if hasattr(st, 'secrets') and key in st.secrets:
                return st.secrets[key]
        except:
            pass
        
        # Fallback to environment variables
        return os.getenv(key)
    
    def _initialize_encryption(self):
        """Initialize encryption with proper key from Streamlit secrets"""
        # Temporarily set environment variables for the DataEncryption class
        encryption_key = self._get_secret('ENCRYPTION_KEY')
        encryption_password = self._get_secret('ENCRYPTION_PASSWORD')
        
        # Store original env vars to restore later
        original_key = os.environ.get('ENCRYPTION_KEY')
        original_password = os.environ.get('ENCRYPTION_PASSWORD')
        
        try:
            # Set the secrets as environment variables temporarily
            if encryption_key:
                os.environ['ENCRYPTION_KEY'] = encryption_key
            if encryption_password:
                os.environ['ENCRYPTION_PASSWORD'] = encryption_password
            
            # Create the encryption instance
            encryption = DataEncryption()
            
        finally:
            # Restore original environment variables
            if original_key is not None:
                os.environ['ENCRYPTION_KEY'] = original_key
            elif 'ENCRYPTION_KEY' in os.environ:
                del os.environ['ENCRYPTION_KEY']
            
            if original_password is not None:
                os.environ['ENCRYPTION_PASSWORD'] = original_password
            elif 'ENCRYPTION_PASSWORD' in os.environ:
                del os.environ['ENCRYPTION_PASSWORD']
        
        return encryption
    
    def _get_file_from_github(self, file_path):
        """Get file content from GitHub"""
        url = f"{self.base_url}/{file_path}"
        response = requests.get(url, headers=self.headers)
        
        if response.status_code == 200:
            file_data = response.json()
            # Decode base64 content
            content = base64.b64decode(file_data['content']).decode('utf-8')
            return content, file_data['sha']
        elif response.status_code == 404:
            return None, None
        else:
            response.raise_for_status()
    
    def _save_file_to_github(self, file_path, content, message, sha=None):
        """Save file content to GitHub"""
        url = f"{self.base_url}/{file_path}"
        
        # Encode content as base64
        encoded_content = base64.b64encode(content.encode('utf-8')).decode('utf-8')
        
        data = {
            'message': message,
            'content': encoded_content,
            'branch': self.branch
        }
        
        # If file exists, include SHA for update
        if sha:
            data['sha'] = sha
        
        response = requests.put(url, headers=self.headers, json=data)
        
        if response.status_code in [200, 201]:
            return response.json()
        else:
            response.raise_for_status()
    
    def load_fixtures(self, week_num):
        """Load fixtures for a specific week from GitHub"""
        file_path = f"fixtures/week{week_num}.csv"
        try:
            content, _ = self._get_file_from_github(file_path)
            if content:
                # Convert CSV string to DataFrame
                from io import StringIO
                return pd.read_csv(StringIO(content))
            return None
        except Exception as e:
            st.error(f"Error loading fixtures for week {week_num}: {e}")
            return None
    
    def load_results(self, week_num):
        """Load results for a specific week from GitHub"""
        file_path = f"results/week{week_num}.csv"
        try:
            content, _ = self._get_file_from_github(file_path)
            if content:
                from io import StringIO
                # Handle both comma and tab-separated files
                # First try comma separator
                try:
                    df = pd.read_csv(StringIO(content))
                except:
                    # If that fails, try tab separator
                    df = pd.read_csv(StringIO(content), sep='\t')
                
                # Clean up column names (remove extra whitespace)
                df.columns = df.columns.str.strip()
                
                # Ensure we have the expected columns
                expected_cols = ['home_team', 'away_team', 'home_score', 'away_score']
                if not all(col in df.columns for col in expected_cols):
                    st.error(f"Week {week_num} results file missing required columns. Found: {list(df.columns)}, Expected: {expected_cols}")
                    return None
                
                return df
            return None
        except Exception as e:
            st.error(f"Error loading results for week {week_num}: {e}")
            return None
    
    def save_predictions(self, username, week_num, predictions):
        """Save user predictions for a week to GitHub (encrypted)"""
        file_path = f"predictions/week{week_num}.json"
        
        try:
            # Load existing predictions for this week
            content, sha = self._get_file_from_github(file_path)
            if content:
                # Decrypt existing data
                try:
                    all_predictions = self.encryption.decrypt_data(content)
                    if all_predictions is None:
                        st.warning("Failed to decrypt existing predictions. Creating new file.")
                        all_predictions = {}
                except Exception as e:
                    st.warning(f"Error decrypting existing predictions: {e}. Creating new file.")
                    all_predictions = {}
            else:
                all_predictions = {}
            
            # Update predictions for this user
            all_predictions[username] = {
                "predictions": predictions,
                "submitted_at": datetime.now().isoformat()
            }
            
            # Encrypt and save back to GitHub
            encrypted_content = self.encryption.encrypt_data(all_predictions)
            if encrypted_content is None:
                st.error("Failed to encrypt predictions data")
                return False
            
            commit_message = f"Update predictions for {username} - Week {week_num}"
            self._save_file_to_github(file_path, encrypted_content, commit_message, sha)
            
            return True
            
        except Exception as e:
            st.error(f"Error saving predictions for {username}, week {week_num}: {e}")
            return False
    
    def load_predictions(self, week_num, username=None):
        """Load predictions for a week from GitHub (decrypt automatically)"""
        file_path = f"predictions/week{week_num}.json"
        try:
            content, _ = self._get_file_from_github(file_path)
            if content:
                try:
                    all_predictions = self.encryption.decrypt_data(content)
                    if all_predictions is None:
                        st.error(f"Failed to decrypt predictions for week {week_num}. Check encryption keys.")
                        return {} if not username else []
                except Exception as e:
                    st.error(f"Error decrypting predictions for week {week_num}: {e}")
                    return {} if not username else []
            else:
                all_predictions = {}
            
            if username:
                user_data = all_predictions.get(username, {})
                return user_data.get("predictions", [])
            return all_predictions
            
        except Exception as e:
            st.error(f"Error loading predictions for week {week_num}: {e}")
            return {} if not username else []
    
    def calculate_points(self, prediction, actual_result):
        """Calculate points for a single match prediction"""
        # Handle pandas Series properly
        if actual_result is None:
            return 0
            
        # Convert pandas Series to dict if needed
        if hasattr(actual_result, 'to_dict'):
            actual_result = actual_result.to_dict()
        
        # Check if we have valid scores
        home_score = actual_result.get('home_score')
        away_score = actual_result.get('away_score')
        
        if home_score is None or away_score is None or pd.isna(home_score) or pd.isna(away_score):
            return 0
        
        # Get prediction scores
        pred_home = prediction.get('home_score', 0)
        pred_away = prediction.get('away_score', 0)
        
        try:
            actual_home = int(float(home_score))
            actual_away = int(float(away_score))
        except (ValueError, TypeError):
            return 0
        
        points = 0
        
        # Exact score
        if pred_home == actual_home and pred_away == actual_away:
            return POINTS_EXACT_SCORE
        
        # Correct result (win/draw/loss)
        pred_result = "draw" if pred_home == pred_away else ("home" if pred_home > pred_away else "away")
        actual_result_outcome = "draw" if actual_home == actual_away else ("home" if actual_home > actual_away else "away")
        
        if pred_result == actual_result_outcome:
            points += POINTS_CORRECT_RESULT
        
        # Goal difference
        pred_diff = pred_home - pred_away
        actual_diff = actual_home - actual_away
        if pred_diff == actual_diff:
            points += POINTS_GOAL_DIFFERENCE
        
        return points
    
    def get_leaderboard(self):
        """Get leaderboard sorted by total points"""
        user_scores = self.calculate_user_scores()
        current_week = self.config.get_current_week()
        
        # Convert to list and sort by total points
        leaderboard = []
        for username, data in user_scores.items():
            avg_points = data["total_points"] / max(data["weeks_played"], 1)
            
            # Get current week points (if available)
            current_week_points = 0
            if current_week > 1:  # Only show current week points if we're past week 1
                previous_week = current_week - 1
                current_week_points = data["weekly_breakdown"].get(f"week_{previous_week}", 0)
            
            leaderboard.append({
                "username": username,
                "display_name": data["display_name"],
                "total_points": data["total_points"],
                "weeks_played": data["weeks_played"],
                "average_points": round(avg_points, 2),
                "current_week_points": current_week_points,
                "weekly_breakdown": data["weekly_breakdown"],
                "manual_adjustments": data.get("manual_adjustments", 0)
            })
        
        return sorted(leaderboard, key=lambda x: x["total_points"], reverse=True)
    
    def has_user_predicted(self, username, week_num):
        """Check if user has already made predictions for a week"""
        predictions = self.load_predictions(week_num, username)
        return len(predictions) > 0
    
    def get_user_predictions_for_week(self, username, week_num):
        """Get user's predictions for a specific week with fixtures"""
        try:
            predictions = self.load_predictions(week_num, username)
            fixtures = self.load_fixtures(week_num)
            
            if not predictions or fixtures is None:
                return None
            
            # Combine predictions with fixtures
            result = []
            for i, (_, fixture) in enumerate(fixtures.iterrows()):
                if i < len(predictions):
                    result.append({
                        "home_team": fixture['home_team'],
                        "away_team": fixture['away_team'],
                        "predicted_home_score": predictions[i]['home_score'],
                        "predicted_away_score": predictions[i]['away_score']
                    })
                    
            return result
        except Exception as e:
            print(f"Error getting user predictions for week {week_num}: {e}")
            return None
    
    def save_fixtures(self, week_num, fixtures_df):
        """Save fixtures for a week to GitHub"""
        file_path = f"fixtures/week{week_num}.csv"
        try:
            csv_content = fixtures_df.to_csv(index=False)
            commit_message = f"Add fixtures for Week {week_num}"
            
            # Check if file already exists
            _, sha = self._get_file_from_github(file_path)
            if sha:
                commit_message = f"Update fixtures for Week {week_num}"
            
            self._save_file_to_github(file_path, csv_content, commit_message, sha)
            return True
        except Exception as e:
            st.error(f"Error saving fixtures for week {week_num}: {e}")
            return False
    
    def save_results(self, week_num, results_df):
        """Save results for a week to GitHub"""
        file_path = f"results/week{week_num}.csv"
        try:
            csv_content = results_df.to_csv(index=False)
            commit_message = f"Add results for Week {week_num}"
            
            # Check if file already exists
            _, sha = self._get_file_from_github(file_path)
            if sha:
                commit_message = f"Update results for Week {week_num}"
                print(f"Updating existing file with SHA: {sha}")
            else:
                print(f"Creating new file: {file_path}")
            
            print(f"Saving to GitHub: {file_path}")
            print(f"Content preview: {csv_content[:200]}...")
            
            self._save_file_to_github(file_path, csv_content, commit_message, sha)
            print(f"Successfully saved results for week {week_num}")
            return True
        except Exception as e:
            print(f"Error saving results for week {week_num}: {e}")
            import traceback
            print(traceback.format_exc())
            if hasattr(st, 'error'):
                st.error(f"Error saving results for week {week_num}: {e}")
            return False
    
    def debug_encryption_status(self):
        """Debug method to check encryption setup and test decryption"""
        st.write("### Encryption Debug Info")
        
        # Check if secrets are available
        encryption_key = self._get_secret('ENCRYPTION_KEY')
        encryption_password = self._get_secret('ENCRYPTION_PASSWORD')
        
        st.write(f"- ENCRYPTION_KEY: {'✅ Set' if encryption_key else '❌ Not set'}")
        st.write(f"- ENCRYPTION_PASSWORD: {'✅ Set' if encryption_password else '❌ Not set (using default)'}")
        
        # Test encryption round-trip
        test_data = {"test": "data", "number": 123}
        encrypted = self.encryption.encrypt_data(test_data)
        if encrypted:
            decrypted = self.encryption.decrypt_data(encrypted)
            if decrypted == test_data:
                st.write("- Encryption test: ✅ Passed")
            else:
                st.write("- Encryption test: ❌ Failed - decryption mismatch")
        else:
            st.write("- Encryption test: ❌ Failed - encryption returned None")
        
        return encryption_key, encryption_password
    
    def save_manual_adjustment(self, username, points_change, reason, admin_user):
        """Save a manual score adjustment to a log file"""
        adjustment_file = "manual_adjustments.json"
        
        try:
            # Load existing adjustments
            content, sha = self._get_file_from_github(adjustment_file)
            if content:
                adjustments = json.loads(content)
            else:
                adjustments = []
            
            # Add new adjustment
            adjustment = {
                "username": username,
                "points_change": points_change,
                "reason": reason,
                "admin_user": admin_user,
                "timestamp": datetime.now().isoformat()
            }
            
            adjustments.append(adjustment)
            
            # Save back to GitHub
            commit_message = f"Manual score adjustment: {username} {'+' if points_change > 0 else ''}{points_change} points"
            self._save_file_to_github(adjustment_file, json.dumps(adjustments, indent=2), commit_message, sha)
            
            return True
            
        except Exception as e:
            st.error(f"Error saving manual adjustment: {e}")
            return False
    
    def get_manual_adjustments(self, username=None):
        """Get manual score adjustments for a user or all users"""
        adjustment_file = "manual_adjustments.json"
        
        try:
            content, _ = self._get_file_from_github(adjustment_file)
            if content:
                adjustments = json.loads(content)
                if username:
                    return [adj for adj in adjustments if adj['username'] == username]
                return adjustments
            return []
        except Exception as e:
            return []
    
    def calculate_user_scores(self):
        """Calculate scores for all users across all completed weeks, including manual adjustments"""
        users = self.config.get_users()
        user_scores = {}
        
        # Initialize scores
        for username in users:
            if username != "admin":  # Skip admin from leaderboard
                user_scores[username] = {
                    "display_name": users[username]["display_name"],
                    "total_points": 0,
                    "weeks_played": 0,
                    "weekly_breakdown": {},
                    "manual_adjustments": 0
                }
        
        current_week = self.config.get_current_week()
        
        # Calculate points for each completed week (only previous weeks, not current week)
        for week in range(1, current_week):
            results = self.load_results(week)
            if results is None or len(results) == 0:
                continue  # Skip weeks without results
            
            predictions = self.load_predictions(week)
            if not predictions:
                continue  # Skip if no predictions found
            
            for username in user_scores:
                if username in predictions:
                    week_points = 0
                    user_data = predictions[username]
                    
                    # Handle both old and new prediction formats
                    if isinstance(user_data, dict) and "predictions" in user_data:
                        user_predictions = user_data["predictions"]
                    elif isinstance(user_data, list):
                        user_predictions = user_data
                    else:
                        continue
                    
                    # Calculate points for each match
                    for i, result_row in results.iterrows():
                        if i < len(user_predictions):
                            points = self.calculate_points(user_predictions[i], result_row)
                            week_points += points
                    
                    user_scores[username]["total_points"] += week_points
                    user_scores[username]["weeks_played"] += 1
                    user_scores[username]["weekly_breakdown"][f"week_{week}"] = week_points
        
        # Add manual adjustments
        all_adjustments = self.get_manual_adjustments()
        for adjustment in all_adjustments:
            username = adjustment['username']
            if username in user_scores:
                user_scores[username]["total_points"] += adjustment['points_change']
                user_scores[username]["manual_adjustments"] += adjustment['points_change']
        
        return user_scores
    
    def export_predictions_to_excel(self, week_num):
        """Export all predictions for a week to an Excel file"""
        try:
            # Load predictions and fixtures
            predictions = self.load_predictions(week_num)
            fixtures = self.load_fixtures(week_num)
            
            if not predictions or fixtures is None:
                return None
            
            # Create a workbook and worksheet
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Week {week_num} Predictions"
            
            # Headers
            headers = ["Match", "Home Team", "Away Team"]
            
            # Get all users who made predictions
            users = list(predictions.keys())
            if "admin" in users:
                users.remove("admin")  # Remove admin from export
            
            # Add user columns
            for username in users:
                user_info = self.config.get_users().get(username, {})
                display_name = user_info.get("display_name", username)
                headers.append(display_name)
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write match data
            for match_idx, (_, fixture) in enumerate(fixtures.iterrows()):
                row = match_idx + 2
                
                # Match info
                ws.cell(row=row, column=1, value=match_idx + 1)
                ws.cell(row=row, column=2, value=fixture['home_team'])
                ws.cell(row=row, column=3, value=fixture['away_team'])
                
                # User predictions
                for user_idx, username in enumerate(users):
                    col = user_idx + 4
                    
                    user_data = predictions.get(username, {})
                    if isinstance(user_data, dict) and "predictions" in user_data:
                        user_predictions = user_data["predictions"]
                    elif isinstance(user_data, list):
                        user_predictions = user_data
                    else:
                        user_predictions = []
                    
                    if match_idx < len(user_predictions):
                        pred = user_predictions[match_idx]
                        prediction_text = f"{pred.get('home_score', 0)}-{pred.get('away_score', 0)}"
                        ws.cell(row=row, column=col, value=prediction_text)
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    else:
                        ws.cell(row=row, column=col, value="No prediction")
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            return wb
            
        except Exception as e:
            st.error(f"Error creating Excel export: {e}")
            return None

# Keep the original class name for backward compatibility
class DataManager(GitHubDataManager):
    pass
